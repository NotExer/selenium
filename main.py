from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from openpyxl import load_workbook
import time
import traceback

# 1️⃣ Inicia el navegador Edge
def iniciar_driver():
    options = Options()
    options.add_argument("--start-maximized")
    driver = webdriver.Edge(options=options)
    return driver

# 2️⃣ Espera a que cargue la página de búsqueda
def esperar_pagina_busqueda(driver):
    wait = WebDriverWait(driver, 10)
    try:
        wait.until(EC.presence_of_element_located((By.ID, "search")))
        print("✅ Página de búsqueda lista.")
    except TimeoutException:
        print("⛔ La página de búsqueda no cargó correctamente.")
        raise
    
    
    
    


# 3️⃣ Busca la empresa y hace clic en el botón buscar
def buscar_empresa(driver, nombre_empresa):
    wait = WebDriverWait(driver, 10)
    try:
        esperar_pagina_busqueda(driver)

        search_box = wait.until(EC.presence_of_element_located((By.ID, "search")))
        search_box.clear()

        # Limpiar cualquier valor residual
        for _ in range(3):
            search_box.send_keys(u'\ue009' + "a")  # Ctrl + A
            search_box.send_keys(u'\ue003')        # Delete
            time.sleep(0.2)
            if search_box.get_attribute("value") == "":
                break

        search_box.send_keys(nombre_empresa)
        print(f"✅ Ingresando el nombre de la empresa: {nombre_empresa}")

        # Buscamos los botones según la vista
        btn_busqueda_mobile = driver.find_elements(By.ID, "btn-busqueda")
        btn_busqueda_desktop_list = driver.find_elements(
            By.XPATH, "//button[contains(@class, 'btn-busqueda') and contains(@class, 'd-sm-block')]"
        )

        if btn_busqueda_mobile and btn_busqueda_mobile[0].is_displayed():
            btn_busqueda_mobile[0].click()
            print("✅ Clic en el botón de búsqueda (MOBILE)")
        elif btn_busqueda_desktop_list and btn_busqueda_desktop_list[0].is_displayed():
            btn_busqueda_desktop_list[0].click()
            print("✅ Clic en el botón de búsqueda (DESKTOP)")
        else:
            print("⛔ No encontré ningún botón visible para buscar.")
            return False

        # Espera para cargar resultados
        time.sleep(2)
        return clic_primer_ver_informacion(driver)

    except Exception as e:
        print(f"⛔ Error en buscar_empresa: {e}")
        traceback.print_exc()
        return False

# 4️⃣ Hace clic en el primer "Ver información" válido
def clic_primer_ver_informacion(driver):
    wait = WebDriverWait(driver, 10)
    try:
        botones_ver_info = wait.until(EC.presence_of_all_elements_located((By.XPATH,
            "//a[contains(text(),'Ver información') and not(contains(., 'Registro de Proponentes'))]"
        )))

        if botones_ver_info:
            botones_ver_info[0].click()
            print("✅ Clic en el primer botón 'Ver información'")
        else:
            print("⛔ No se encontraron botones 'Ver información'.")
            return False

        # Esperamos la pestaña de actividad económica
        time.sleep(1)
        actividad_economica_tab = wait.until(EC.element_to_be_clickable((By.ID, "detail-tabs-tab-pestana_economica")))
        actividad_economica_tab.click()
        print("✅ Clic en la pestaña 'Actividad económica'")
        time.sleep(1)
        return True

    except TimeoutException:
        print("⛔ Timeout en 'clic_primer_ver_informacion'.")
        return False
    except Exception as e:
        print(f"⛔ Error en clic_primer_ver_informacion: {e}")
        traceback.print_exc()
        return False

# 5️⃣ Extrae el CIIU y la descripción
def obtener_actividad_economica(driver):
    wait = WebDriverWait(driver, 10)
    try:
        pestana = wait.until(EC.presence_of_element_located((By.ID, "detail-tabs-tabpane-pestana_economica")))
        registros = pestana.find_elements(By.CLASS_NAME, "registroapi")

        if not registros:
            print("⛔ No hay registros en Actividad económica.")
            return "NO DISPONIBLE", "NO DISPONIBLE"

        primer_registro = registros[0]

        codigo_element = primer_registro.find_element(By.CSS_SELECTOR, "p.registroapi__etiqueta.font-rues-small")
        codigo_ciiu = codigo_element.text.strip()

        descripcion_element = primer_registro.find_element(By.CSS_SELECTOR, "p.registroapi__valor")
        descripcion = descripcion_element.text.strip()

        print(f"✅ Código CIIU: {codigo_ciiu} - Descripción: {descripcion}")
        return codigo_ciiu, descripcion

    except TimeoutException:
        print("⛔ Timeout esperando los datos de Actividad económica.")
        return "NO DISPONIBLE", "NO DISPONIBLE"
    except Exception as e:
        print(f"⛔ Error al obtener actividad económica: {e}")
        traceback.print_exc()
        return "NO DISPONIBLE", "NO DISPONIBLE"

# 6️⃣ Clic en el logo para volver a la búsqueda
def volver_a_busqueda(driver):
    wait = WebDriverWait(driver, 10)
    try:
        logo = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "navbar-brand")))
        logo.click()
        print("✅ Clic en el logo para volver a búsqueda")

        time.sleep(2)  # Espera 2 segundos después de hacer clic en el logo
        
        # Cierra el modal inicial si aparece otra vez
        try:
            cerrar_modal = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "swal2-close")))
            cerrar_modal.click()
            print("✅ Modal cerrado nuevamente después de volver a búsqueda")
        except TimeoutException:
            print("ℹ️ No apareció el modal después de volver a búsqueda.")
        
        esperar_pagina_busqueda(driver)

    except TimeoutException:
        print("⛔ No encontré el logo para volver a la búsqueda.")
    except Exception as e:
        print(f"⛔ Error al volver a búsqueda: {e}")
        traceback.print_exc()



# 7️⃣ Lee el Excel, ejecuta el proceso y actualiza las columnas U y V
def ejecutar_busqueda_excel(ruta_excel):
    driver = None
    try:
        workbook = load_workbook(filename=ruta_excel)
        hoja = workbook['Hoja3']

        encabezados = [cell.value for cell in next(hoja.iter_rows(min_row=1, max_row=1))]

        try:
            col_razon_social = encabezados.index('RAZON SOCIAL')
        except ValueError:
            print("⛔ ERROR: No se encuentra la columna 'RAZON SOCIAL'.")
            return

        driver = iniciar_driver()
        driver.get("https://ruesfront.rues.org.co")

        wait = WebDriverWait(driver, 10)

        # Cierra el modal inicial si aparece
        try:
            cerrar_modal = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "swal2-close")))
            cerrar_modal.click()
            print("✅ Modal cerrado exitosamente")
        except TimeoutException:
            print("ℹ️ No apareció el modal de inicio.")

        esperar_pagina_busqueda(driver)

        for index, fila in enumerate(hoja.iter_rows(min_row=2), start=2):
            razon_social = fila[col_razon_social].value

            if not razon_social:
                print(f"⚠️ Fila {index} sin razón social. Saltando...")
                continue

            print(f"\n➡️ Procesando empresa (fila {index}): {razon_social}")

            exito = buscar_empresa(driver, razon_social)
            if not exito:
                print(f"⚠️ No se pudo procesar la empresa en fila {index}. Saltando...")
                volver_a_busqueda(driver)
                continue

            codigo_ciiu, descripcion = obtener_actividad_economica(driver)

            hoja.cell(row=index, column=21, value=codigo_ciiu)
            hoja.cell(row=index, column=22, value=descripcion)

            volver_a_busqueda(driver)

        workbook.save(ruta_excel)
        print("\n✅ Proceso finalizado. Excel actualizado y guardado.")

    except Exception as e:
        print(f"⛔ Error en ejecutar_busqueda_excel: {e}")
        traceback.print_exc()

    finally:
        if driver:
            driver.quit()
            print("✅ Navegador cerrado.")

# 8️⃣ Ejecuta el script
if __name__ == "__main__":
    ruta_excel = "empresas.xlsx"
    ejecutar_busqueda_excel(ruta_excel)
